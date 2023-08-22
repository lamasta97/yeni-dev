import json
import openpyxl

# Giriş ve çıkış dosya adları
input_file = "input.xlsx"
output_file = "output.xlsx"

# Giriş Excel dosyasını aç
input_workbook = openpyxl.load_workbook(input_file)
input_sheet = input_workbook.active

# Çıktı verilerini tutacak liste
output_data = []

# Tüm hücreleri gez ve verileri JSON olarak yükle
for row in input_sheet.iter_rows():
    for cell in row:
        value = cell.value
        if value:
            try:
                data = json.loads("{" + value + "}")  # JSON olarak yükle
                if isinstance(data, dict):
                    for key in data.keys():
                        output_data.append(key.strip())  # Sadece sol tarafları al
            except json.JSONDecodeError:
                pass  # Hatalı JSON geç, devam et

# Yeni bir Excel dosyası oluştur ve verileri yaz
output_workbook = openpyxl.Workbook()
output_sheet = output_workbook.active
for data in output_data:
    output_sheet.append([data])

# Çıkış Excel dosyasını kaydet
output_workbook.save(output_file)

print("İşlem tamamlandı.")
